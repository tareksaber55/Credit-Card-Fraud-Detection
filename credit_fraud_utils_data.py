import joblib
from matplotlib import pyplot as plt
import numpy as np
import pandas as pd
from sklearn.metrics import classification_report, confusion_matrix, precision_recall_curve,average_precision_score,f1_score
from openpyxl import Workbook, load_workbook
import os

def load_data(train_path,val_path,test_path,args=None):
    df_train = pd.read_csv(train_path)
    df_val =  pd.read_csv(val_path)
    df_test = pd.read_csv(test_path)

    # we cannot work with time as it is , we should convert it to Hours of day 
    for df in [df_train, df_val, df_test]:
        df['Hour'] = df['Time'] // 3600 % 24

    if(args and args.outliers_features):
        df_train = del_outliers(df_train,args.outliers_features,args.outliers_factor)

    x_train = df_train.drop(['Time','Class'],axis=1)
    t_train = df_train['Class']
    
    x_val = df_val.drop(['Time','Class'],axis=1)
    t_val = df_val['Class']

    x_test = df_test.drop(['Time','Class'],axis=1)
    t_test = df_test['Class'] 

    return x_train,t_train,x_val,t_val,x_test,t_test




def show_distributions(df):

    fig,axs = plt.subplots(6,5,figsize=(15,12))

    for i,ax in enumerate(axs.flat):
        feature = df.columns[i]
        ax.boxplot([df[df.Class == 0][feature] ,
                   df[df.Class == 1][feature]])
        
        ax.set_title(feature)
        ax.set_xticks([1,2])
        ax.set_xticklabels(['No Fraud','Fraud'])
    plt.suptitle('Distribution of Features Per Class')
    plt.tight_layout()
    plt.show()



def del_outliers(df,cols,factor):
    """
    Remove outliers only from Class == 0 using the IQR method.
    """
    df_0 = df[df.Class == 0]
    df_1 = df[df.Class == 1]
    for col in cols:
        q1 = df_0[col].quantile(0.25)
        q3 = df_0[col].quantile(0.75)
        iqr = q3 - q1
        lower = q1 - factor*iqr
        upper = q3 + factor*iqr
        df_0 = df_0[(df_0[col] <= upper) & (df_0[col] >= lower)]
    df = pd.concat([df_0,df_1],axis=0).reset_index(drop=True)
    df = df.sample(frac=1, random_state=42).reset_index(drop=True) # shuffle
    return df




def bestthreshold(model,x,t,plot=True):
    t_prob = model.predict_proba(x)[:,1]
    precision,recall,threshold = precision_recall_curve(t,t_prob)
    precision = precision[:-1]
    recall = recall[:-1]
    f1_score = (2*precision*recall)/(recall+precision)
    if plot:
        plt.plot(threshold,f1_score)
        plt.xlabel('Thresholds')
        plt.ylabel('F1-Scores')
        plt.title('F1-Score Results')
        plt.show()

    max_idx = np.argmax(f1_score)
    return threshold[max_idx]

def bestthreshold_recall_with_precision(model, x, t, min_precision=0.5, plot=True):
    t_prob = model.predict_proba(x)[:, 1]
    precision, recall, threshold = precision_recall_curve(t, t_prob)

    # Remove last value (because the last precision/recall value has no threshold)
    precision = precision[:-1]
    recall = recall[:-1]

    # Filter thresholds that satisfy the precision constraint
    valid_indices = np.where(precision >= min_precision)[0]

    # If no threshold satisfies the precision requirement
    if len(valid_indices) == 0:
        print(f"No threshold satisfies precision ≥ {min_precision}")
        return None

    # Among valid thresholds, pick threshold with maximum recall
    best_idx = valid_indices[np.argmax(recall[valid_indices])]

    if plot:
        plt.figure(figsize=(7,5))
        plt.plot(threshold, recall, label='Recall')
        plt.plot(threshold, precision, label='Precision')
        plt.axhline(min_precision, color='red', linestyle='--', label='Precision Constraint')
        plt.xlabel('Threshold')
        plt.ylabel('Score')
        plt.title('Precision & Recall vs Threshold')
        plt.legend()
        plt.show()

    return threshold[best_idx]




def report(model,x,t, threshold=0.2):
    t_prob = model.predict_proba(x)[:,1]
    t_pred = (t_prob>=threshold).astype(int)
    cm = confusion_matrix(t,t_pred)
    cr = classification_report(t,t_pred)
    f1 = f1_score(t,t_pred)
    avgp = average_precision_score(t,t_prob)
    print("Confusion Matrix:")
    print(cm)
    print("\nClassification Report:")
    print(cr)
    print(f"F1-Score: {f1:.4f}")
    print(f'average_precision: {avgp}')
    return f1,avgp



def save_results_to_excel(args,params,f1, avg_precision,threshold,
                          file_path="results/model_results.xlsx"):
    """
    Save model training results into an Excel file inside /results folder.

    Creates the folder if it does not exist.
    Appends rows if the Excel file already exists.
    """

    # Ensure results directory exists
    os.makedirs(os.path.dirname(file_path), exist_ok=True)

    # If file doesn’t exist → create new workbook
    if not os.path.exists(file_path):
        wb = Workbook()
        ws = wb.active
        ws.title = "Models Results"

        # Header
        ws.append([
            "Model","Scaler", "Outlier Features","Sampling Technique","Sampling Factor",
            "Parameters", "Threshold" ,"F1-Score", "Average Precision"
        ])

        wb.save(file_path)

    # Load workbook again to append
    wb = load_workbook(file_path)
    ws = wb["Models Results"]

    # Append results row
    ws.append([
        args.model,
        args.scaler,
        str(args.outliers_features) + f' With factor {args.outliers_factor}' if args.outliers_features else "None",
        args.sampler,
        args.factor if args.sampler != 'None' else 0,
        str(params),
        threshold,
        f1,
        avg_precision
    ])

    wb.save(file_path)
    print(f"[INFO] Results saved to {file_path}")

def save_model(model,threshold):
    model_dictionary = {
        'model' : model,
        'threshold' : threshold
    }
    joblib.dump(model_dictionary, 'model_dictionary.joblib')